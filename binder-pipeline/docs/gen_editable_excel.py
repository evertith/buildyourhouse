#!/usr/bin/env python3
"""Editable Excel (.xlsx) companions for the Owner-Builder Job Site Binder.

Four working workbooks, built with real formulas rather than printed blanks:

  1.2-budget-tracking-spreadsheet.xlsx  19 cost categories, per-category
                                        subtotals, variance, grand total
  7.1-expense-tracking.xlsx             expense ledger with running total and a
                                        SUMIF-driven category summary
  7.4-payment-tracking.xlsx             draw/invoice ledger with balance owed
  8.3-material-calculators.xlsx         five calculators, yellow input cells

Output: out/editable-documents/excel/
"""

import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Protection, Side
from openpyxl.utils import get_column_letter
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.datavalidation import DataValidation

HERE = os.path.dirname(os.path.abspath(__file__))
PIPELINE = os.path.dirname(HERE)
OUT_DIR = os.path.join(PIPELINE, "out", "editable-documents", "excel")

# -- house style ------------------------------------------------------------

BODY_FONT = "Calibri"
COPYRIGHT = "© 2026 Build Your House · build-your-house.com"

HEADER_FILL = PatternFill("solid", fgColor="E6E6E6")
SUBTOTAL_FILL = PatternFill("solid", fgColor="F2F2F2")
CATEGORY_FILL = PatternFill("solid", fgColor="D9D9D9")
INPUT_FILL = PatternFill("solid", fgColor="FFF6CC")
TOTAL_FILL = PatternFill("solid", fgColor="DCE6F1")

TITLE_FONT = Font(name=BODY_FONT, size=15, bold=True)
SECTION_FONT = Font(name=BODY_FONT, size=11, bold=True)
HEADER_FONT = Font(name=BODY_FONT, size=11, bold=True)
BODY = Font(name=BODY_FONT, size=11)
BOLD = Font(name=BODY_FONT, size=11, bold=True)
NOTE_FONT = Font(name=BODY_FONT, size=9, italic=True, color="595959")

THIN = Side(style="thin", color="BFBFBF")
MED = Side(style="thin", color="808080")
BOX = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)
HEADER_BORDER = Border(left=THIN, right=THIN, top=MED, bottom=MED)

MONEY = '$#,##0.00'
NUMBER = '#,##0.00'
INT = '#,##0'
PCT = '0%'
DATE = 'mm/dd/yyyy'

UNLOCKED = Protection(locked=False)


def style_header_row(ws, row, first_col, last_col, wrap=True):
    for col in range(first_col, last_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = HEADER_BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center",
                                   wrap_text=wrap)
    ws.row_dimensions[row].height = 28


def set_widths(ws, widths):
    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = width


def title_block(ws, title, subtitle=None, span=5):
    ws["A1"] = title
    ws["A1"].font = TITLE_FONT
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=span)
    ws.row_dimensions[1].height = 22
    if subtitle:
        ws["A2"] = subtitle
        ws["A2"].font = NOTE_FONT
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=span)


def footer_note(ws, row, span, text=COPYRIGHT):
    cell = ws.cell(row=row, column=1, value=text)
    cell.font = NOTE_FONT
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)


def finish(wb, filename):
    os.makedirs(OUT_DIR, exist_ok=True)
    path = os.path.join(OUT_DIR, filename)
    wb.properties.creator = "Build Your House"
    wb.properties.title = os.path.splitext(filename)[0]
    wb.save(path)
    return path


# ---------------------------------------------------------------------------
# 1.2  Budget tracking spreadsheet
# ---------------------------------------------------------------------------

BUDGET_CATEGORIES = [
    ("LAND & SITE COSTS", [
        "Land purchase / lot cost",
        "Closing costs and fees",
        "Land survey",
        "Soil / percolation testing",
    ]),
    ("SITE PREPARATION", [
        "Site clearing and tree removal",
        "Grading and excavation",
        "Driveway / access road",
        "Erosion control measures",
        "Temporary utilities (power, water)",
        "Portable toilet rental",
        "Dumpster rental",
        "Storage container / job box",
    ]),
    ("FOUNDATION", [
        "Foundation excavation",
        "Gravel / stone base",
        "Rebar and wire mesh",
        "Concrete (footers)",
        "Concrete (foundation walls / slab)",
        "Concrete pump rental",
        "Form lumber and supplies",
        "Anchor bolts and hardware",
        "Waterproofing / dampproofing",
        "Drainage tile and pipe",
        "Vapor barrier / poly sheeting",
        "Backfill and compaction",
        "Foundation labor (if subbed)",
    ]),
    ("FRAMING", [
        "Framing lumber (studs, plates, joists)",
        "Engineered lumber (LVL, I-joists, etc.)",
        "Roof trusses or rafters",
        "Sheathing (wall, floor, roof)",
        "House wrap / moisture barrier",
        "Nails, screws, hangers, hardware",
        "Framing labor (if subbed)",
        "Crane rental (truss setting)",
    ]),
    ("ROOFING", [
        "Roofing underlayment (felt / synthetic)",
        "Shingles / roofing material",
        "Ridge vent and accessories",
        "Drip edge and flashing",
        "Gutters and downspouts",
        "Roofing labor (if subbed)",
    ]),
    ("EXTERIOR FINISHES", [
        "Siding material",
        "Exterior trim boards",
        "Soffit and fascia",
        "Exterior paint / stain",
        "Exterior doors (front, rear, garage)",
        "Windows",
        "Exterior labor (if subbed)",
    ]),
    ("PLUMBING ROUGH-IN", [
        "Supply pipes (PEX, copper, etc.)",
        "Drain / waste / vent (DWV) pipes",
        "Fittings, connectors, valves",
        "Water heater",
        "Plumbing rough-in labor (if subbed)",
    ]),
    ("HVAC ROUGH-IN", [
        "HVAC unit(s) — furnace / AC / heat pump",
        "Ductwork and registers",
        "Ventilation fans (bath, kitchen)",
        "HVAC labor / installation (if subbed)",
    ]),
    ("ELECTRICAL ROUGH-IN", [
        "Electrical panel / breaker box",
        "Wire and cable (Romex, etc.)",
        "Boxes, connectors, staples",
        "Service entrance / meter base",
        "Electrical rough-in labor (if subbed)",
    ]),
    ("INSULATION", [
        "Wall insulation (fiberglass / spray foam)",
        "Ceiling / attic insulation",
        "Floor / rim joist insulation",
        "Insulation labor (if subbed)",
    ]),
    ("DRYWALL", [
        'Drywall sheets (1/2", 5/8")',
        "Joint compound (mud)",
        "Tape, corner bead, screws",
        "Texture materials (if applicable)",
        "Drywall labor (hang, tape, finish)",
    ]),
    ("DOORS & INTERIOR TRIM", [
        "Interior doors (slab and pre-hung)",
        "Door hardware (knobs, hinges, locks)",
        "Baseboards",
        "Window and door casing",
        "Crown molding (if applicable)",
        "Closet shelving and rods",
        "Trim labor (if subbed)",
    ]),
    ("FLOORING", [
        "LVP / vinyl plank flooring",
        "Tile (bathroom, kitchen, entry)",
        "Hardwood flooring",
        "Carpet (bedrooms, stairs)",
        "Underlayment and adhesives",
        "Flooring labor (if subbed)",
    ]),
    ("CABINETS & COUNTERTOPS", [
        "Kitchen cabinets (base and upper)",
        "Bathroom vanities",
        "Kitchen countertops",
        "Bathroom countertops",
        "Cabinet hardware (pulls, hinges)",
        "Cabinet installation labor (if subbed)",
    ]),
    ("PLUMBING FIXTURES", [
        "Kitchen sink and faucet",
        "Bathroom sinks and faucets",
        "Toilets",
        "Bathtubs",
        "Shower enclosures and doors",
        "Shower / tub fixtures and valves",
        "Plumbing fixture installation labor",
    ]),
    ("ELECTRICAL FIXTURES & FINISH", [
        "Light fixtures (interior and exterior)",
        "Ceiling fans",
        "Receptacles and switches",
        "Cover plates",
        "Doorbell / smoke / CO detectors",
        "Electrical trim-out labor (if subbed)",
    ]),
    ("PAINTING", [
        "Interior primer",
        "Interior paint",
        "Exterior primer and paint",
        "Painting supplies (brushes, rollers, tape)",
        "Painting labor (if subbed)",
    ]),
    ("FINISH ITEMS & APPLIANCES", [
        "Kitchen appliances (range, fridge, DW)",
        "Microwave / range hood",
        "Laundry appliances (washer / dryer)",
        "Mirrors and shower doors",
        "Tile backsplash materials and labor",
        "Garage door and opener",
        "Deck / patio materials and labor",
    ]),
    ("LANDSCAPING & FINAL SITE WORK", [
        "Final grading and drainage",
        "Topsoil and seeding / sod",
        "Trees, shrubs, and plantings",
        "Mulch and landscape fabric",
        "Driveway paving / concrete",
        "Walkways and steps",
        "Mailbox and house numbers",
    ]),
    ("PERMITS, FEES & INSURANCE", [
        "Building permit",
        "Electrical permit",
        "Plumbing permit",
        "Mechanical / HVAC permit",
        "Septic / well permits",
        "Impact / tap fees (water, sewer)",
        "Builder's risk insurance",
        "General liability insurance",
        "Plan review / engineering fees",
    ]),
    ("CONTINGENCY & MISCELLANEOUS", [
        "Contingency fund (10–15% recommended)",
        "Tool purchases / rentals",
        "Miscellaneous supplies",
        "Cleanup and waste removal",
        "Inspection fees (beyond permit costs)",
        "Other",
        "Other",
    ]),
]


def build_budget_tracker():
    wb = Workbook()
    ws = wb.active
    ws.title = "Budget Tracker"
    ws.sheet_view.showGridLines = False

    title_block(ws, "COMPLETE BUDGET TRACKING SPREADSHEET",
                "Enter estimated costs up front, then fill in actual costs as "
                "invoices arrive. Variance = Actual − Estimated; a positive "
                "variance means you are over budget.")

    info = [
        ("Project Name", "Total Budget"),
        ("Owner-Builder", "Date Created"),
        ("Lender / Financing", "Last Updated"),
    ]
    for offset, (left, right) in enumerate(info):
        row = 4 + offset
        ws.cell(row=row, column=1, value=left + ":").font = BOLD
        ws.cell(row=row, column=2).border = Border(bottom=THIN)
        ws.cell(row=row, column=2).fill = INPUT_FILL
        ws.cell(row=row, column=2).protection = UNLOCKED
        ws.cell(row=row, column=4, value=right + ":").font = BOLD
        ws.cell(row=row, column=5).border = Border(bottom=THIN)
        ws.cell(row=row, column=5).fill = INPUT_FILL
        ws.cell(row=row, column=5).protection = UNLOCKED
    ws["E4"].number_format = MONEY  # Total Budget; the summary block reads $E$4

    header_row = 8
    headers = ["Item", "Estimated Cost", "Actual Cost", "Variance (+/−)",
               "Vendor / Notes"]
    for col, text in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=text)
    style_header_row(ws, header_row, 1, 5)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    row = header_row + 1
    subtotal_cells = []
    for category, items in BUDGET_CATEGORIES:
        cat_row = row
        cell = ws.cell(row=cat_row, column=1, value=category)
        cell.font = SECTION_FONT
        for col in range(1, 6):
            ws.cell(row=cat_row, column=col).fill = CATEGORY_FILL
            ws.cell(row=cat_row, column=col).border = BOX
        ws.row_dimensions[cat_row].height = 20
        row += 1

        first_item = row
        for item in items:
            ws.cell(row=row, column=1, value=item).font = BODY
            for col in (2, 3):
                cell = ws.cell(row=row, column=col)
                cell.number_format = MONEY
                cell.fill = INPUT_FILL
                cell.protection = UNLOCKED
            var = ws.cell(row=row, column=4)
            var.value = ('=IF(AND(B{r}="",C{r}=""),"",C{r}-B{r})'.format(r=row))
            var.number_format = MONEY
            notes = ws.cell(row=row, column=5)
            notes.fill = INPUT_FILL
            notes.protection = UNLOCKED
            for col in range(1, 6):
                ws.cell(row=row, column=col).border = BOX
            row += 1
        last_item = row - 1

        sub_row = row
        ws.cell(row=sub_row, column=1,
                value="%s SUBTOTAL" % category).font = BOLD
        for col, letter in ((2, "B"), (3, "C")):
            cell = ws.cell(row=sub_row, column=col)
            cell.value = "=SUM({L}{a}:{L}{b})".format(L=letter, a=first_item,
                                                      b=last_item)
            cell.number_format = MONEY
            cell.font = BOLD
        var = ws.cell(row=sub_row, column=4)
        var.value = "=C{r}-B{r}".format(r=sub_row)
        var.number_format = MONEY
        var.font = BOLD
        for col in range(1, 6):
            ws.cell(row=sub_row, column=col).fill = SUBTOTAL_FILL
            ws.cell(row=sub_row, column=col).border = BOX
        subtotal_cells.append(sub_row)
        row += 2

    # -- grand total & summary --------------------------------------------
    grand_row = row
    ws.cell(row=grand_row, column=1, value="GRAND TOTAL — ALL CATEGORIES").font = \
        Font(name=BODY_FONT, size=12, bold=True)
    for col, letter in ((2, "B"), (3, "C")):
        cell = ws.cell(row=grand_row, column=col)
        cell.value = "=" + "+".join("%s%d" % (letter, r) for r in subtotal_cells)
        cell.number_format = MONEY
        cell.font = Font(name=BODY_FONT, size=12, bold=True)
    var = ws.cell(row=grand_row, column=4)
    var.value = "=C{r}-B{r}".format(r=grand_row)
    var.number_format = MONEY
    var.font = Font(name=BODY_FONT, size=12, bold=True)
    for col in range(1, 6):
        ws.cell(row=grand_row, column=col).fill = TOTAL_FILL
        ws.cell(row=grand_row, column=col).border = Border(
            left=THIN, right=THIN, top=MED, bottom=MED)
    ws.row_dimensions[grand_row].height = 22

    summary = [
        ("Total variance (Actual − Estimated)",
         "=C{g}-B{g}".format(g=grand_row), MONEY),
        ("Remaining budget available",
         '=IF($E$4="","",$E$4-C{g})'.format(g=grand_row), MONEY),
        ("Percentage of total budget used",
         '=IF($E$4="","",C{g}/$E$4)'.format(g=grand_row), PCT),
    ]
    row = grand_row + 2
    for label, formula, fmt in summary:
        ws.cell(row=row, column=1, value=label).font = BOLD
        cell = ws.cell(row=row, column=3, value=formula)
        cell.number_format = fmt
        cell.font = BOLD
        cell.border = BOX
        cell.fill = SUBTOTAL_FILL
        row += 1

    ws.cell(row=row + 1, column=1,
            value="Yellow cells are for your entries. White cells calculate "
                  "automatically — leave them alone.").font = NOTE_FONT
    footer_note(ws, row + 2, 5)

    set_widths(ws, [44, 16, 16, 16, 36])
    ws.page_setup.orientation = "portrait"
    ws.print_title_rows = "%d:%d" % (header_row, header_row)

    return finish(wb, "1.2-budget-tracking-spreadsheet.xlsx")


# ---------------------------------------------------------------------------
# 7.1  Expense tracking
# ---------------------------------------------------------------------------

EXPENSE_CATEGORIES = [
    "Land & Site Costs",
    "Site Preparation",
    "Foundation",
    "Framing",
    "Roofing",
    "Exterior (Siding, Windows, Doors)",
    "Rough-In (Plumbing, Electrical, HVAC)",
    "Insulation",
    "Drywall",
    "Doors & Interior Trim",
    "Flooring",
    "Cabinets & Countertops",
    "Fixtures (Plumbing, Lighting)",
    "Paint & Finishes",
    "Appliances",
    "Landscaping",
    "Permits & Fees",
    "Tools & Equipment",
    "Labor (Subcontractors)",
    "Contingency",
    "Other",
]

PAYMENT_METHODS = [
    "Cash",
    "Check",
    "Debit Card",
    "Credit Card",
    "ACH / Transfer",
    "Draw / Construction Loan",
    "Owner Financing",
]

EXPENSE_ROWS = 60


def build_expense_tracking():
    wb = Workbook()
    ws = wb.active
    ws.title = "Expense Log"
    ws.sheet_view.showGridLines = False

    lists = wb.create_sheet("Lists")
    lists["A1"] = "Budget Category"
    lists["A1"].font = HEADER_FONT
    lists["A1"].fill = HEADER_FILL
    lists["C1"] = "Payment Method"
    lists["C1"].font = HEADER_FONT
    lists["C1"].fill = HEADER_FILL
    for i, name in enumerate(EXPENSE_CATEGORIES, start=2):
        lists.cell(row=i, column=1, value=name).font = BODY
    for i, name in enumerate(PAYMENT_METHODS, start=2):
        lists.cell(row=i, column=3, value=name).font = BODY
    set_widths(lists, [38, 4, 28])
    lists["E1"] = ("Edit these lists to change the dropdown choices on the "
                   "Expense Log sheet.")
    lists["E1"].font = NOTE_FONT
    lists.sheet_view.showGridLines = False

    cat_range = "Lists!$A$2:$A$%d" % (len(EXPENSE_CATEGORIES) + 1)
    pay_range = "Lists!$C$2:$C$%d" % (len(PAYMENT_METHODS) + 1)
    wb.defined_names["ExpenseCategories"] = DefinedName(
        "ExpenseCategories", attr_text=cat_range)
    wb.defined_names["PaymentMethods"] = DefinedName(
        "PaymentMethods", attr_text=pay_range)

    title_block(ws, "EXPENSE TRACKING LOG",
                "Record every purchase the day it happens. File the receipt, "
                "then tick it off here.", span=7)
    ws["A4"] = "Month:"
    ws["A4"].font = BOLD
    ws["B4"].fill = INPUT_FILL
    ws["B4"].border = Border(bottom=THIN)
    ws["B4"].protection = UNLOCKED
    ws["D4"] = "Project:"
    ws["D4"].font = BOLD
    ws["E4"].fill = INPUT_FILL
    ws["E4"].border = Border(bottom=THIN)
    ws["E4"].protection = UNLOCKED

    header_row = 6
    headers = ["Date", "Category", "Vendor", "Description", "Payment Method",
               "Amount", "Running Total"]
    for col, text in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=text)
    style_header_row(ws, header_row, 1, 7)
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    first = header_row + 1
    last = first + EXPENSE_ROWS - 1
    for row in range(first, last + 1):
        ws.cell(row=row, column=1).number_format = DATE
        ws.cell(row=row, column=6).number_format = MONEY
        total = ws.cell(row=row, column=7)
        total.value = ('=IF(F{r}="","",SUM($F${f}:F{r}))'
                       .format(r=row, f=first))
        total.number_format = MONEY
        total.font = BODY
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.protection = UNLOCKED
        for col in range(1, 8):
            ws.cell(row=row, column=col).border = BOX
        ws.row_dimensions[row].height = 18

    cat_dv = DataValidation(type="list", formula1="=ExpenseCategories",
                            allow_blank=True, showDropDown=False)
    cat_dv.error = "Pick a budget category from the list on the Lists sheet."
    cat_dv.errorTitle = "Unknown category"
    ws.add_data_validation(cat_dv)
    cat_dv.add("B%d:B%d" % (first, last))

    pay_dv = DataValidation(type="list", formula1="=PaymentMethods",
                            allow_blank=True, showDropDown=False)
    ws.add_data_validation(pay_dv)
    pay_dv.add("E%d:E%d" % (first, last))

    total_row = last + 1
    ws.cell(row=total_row, column=1, value="TOTAL THIS SHEET").font = BOLD
    cell = ws.cell(row=total_row, column=6,
                   value="=SUM(F%d:F%d)" % (first, last))
    cell.number_format = MONEY
    cell.font = BOLD
    for col in range(1, 8):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = Border(
            left=THIN, right=THIN, top=MED, bottom=MED)

    footer_note(ws, total_row + 2, 7)
    set_widths(ws, [12, 34, 26, 40, 22, 14, 16])
    ws.print_title_rows = "%d:%d" % (header_row, header_row)

    # -- monthly category summary -----------------------------------------
    summary = wb.create_sheet("Category Summary")
    summary.sheet_view.showGridLines = False
    title_block(summary, "CATEGORY SUMMARY",
                "Budgeted amounts are yours to enter; spend is pulled from the "
                "Expense Log automatically.", span=4)
    for col, text in enumerate(
            ["Budget Category", "Budgeted", "Spent (from log)", "Remaining"],
            start=1):
        summary.cell(row=4, column=col, value=text)
    style_header_row(summary, 4, 1, 4)
    summary.freeze_panes = summary.cell(row=5, column=1)

    srow = 5
    for name in EXPENSE_CATEGORIES:
        summary.cell(row=srow, column=1, value=name).font = BODY
        budget = summary.cell(row=srow, column=2)
        budget.number_format = MONEY
        budget.fill = INPUT_FILL
        budget.protection = UNLOCKED
        spent = summary.cell(row=srow, column=3)
        spent.value = ("=SUMIF('Expense Log'!$B${f}:$B${l},$A{r},"
                       "'Expense Log'!$F${f}:$F${l})"
                       .format(f=first, l=last, r=srow))
        spent.number_format = MONEY
        remaining = summary.cell(row=srow, column=4)
        remaining.value = '=IF(B{r}="","",B{r}-C{r})'.format(r=srow)
        remaining.number_format = MONEY
        for col in range(1, 5):
            summary.cell(row=srow, column=col).border = BOX
        srow += 1

    summary.cell(row=srow, column=1, value="TOTAL").font = BOLD
    for col, letter in ((2, "B"), (3, "C"), (4, "D")):
        cell = summary.cell(row=srow, column=col)
        cell.value = "=SUM({L}5:{L}{last})".format(L=letter, last=srow - 1)
        cell.number_format = MONEY
        cell.font = BOLD
    for col in range(1, 5):
        summary.cell(row=srow, column=col).fill = TOTAL_FILL
        summary.cell(row=srow, column=col).border = Border(
            left=THIN, right=THIN, top=MED, bottom=MED)

    footer_note(summary, srow + 2, 4)
    set_widths(summary, [40, 18, 20, 18])

    return finish(wb, "7.1-expense-tracking.xlsx")


# ---------------------------------------------------------------------------
# 7.4  Payment tracking
# ---------------------------------------------------------------------------

PAYMENT_ROWS = 40


def build_payment_tracking():
    wb = Workbook()
    ws = wb.active
    ws.title = "Payment Tracking"
    ws.sheet_view.showGridLines = False

    lists = wb.create_sheet("Lists")
    lists["A1"] = "Lien Waiver Received?"
    lists["A1"].font = HEADER_FONT
    lists["A1"].fill = HEADER_FILL
    for i, value in enumerate(["Yes", "No", "N/A"], start=2):
        lists.cell(row=i, column=1, value=value).font = BODY
    set_widths(lists, [26])
    lists.sheet_view.showGridLines = False
    wb.defined_names["WaiverStatus"] = DefinedName(
        "WaiverStatus", attr_text="Lists!$A$2:$A$4")

    title_block(ws, "SUBCONTRACTOR PAYMENT TRACKING",
                "Never release a payment without the matching lien waiver in "
                "hand. Log the draw the day the check leaves your desk.", span=9)

    ws["A4"] = "Project:"
    ws["A4"].font = BOLD
    ws["B4"].fill = INPUT_FILL
    ws["B4"].border = Border(bottom=THIN)
    ws["B4"].protection = UNLOCKED

    header_row = 6
    headers = ["Date", "Payee / Sub", "Draw / Invoice #", "Amount Due",
               "Amount Paid", "Balance", "Check / Ref #",
               "Lien Waiver Received?", "Notes"]
    for col, text in enumerate(headers, start=1):
        ws.cell(row=header_row, column=col, value=text)
    style_header_row(ws, header_row, 1, len(headers))
    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)

    first = header_row + 1
    last = first + PAYMENT_ROWS - 1
    for row in range(first, last + 1):
        ws.cell(row=row, column=1).number_format = DATE
        for col in (4, 5):
            ws.cell(row=row, column=col).number_format = MONEY
        balance = ws.cell(row=row, column=6)
        balance.value = ('=IF(AND(D{r}="",E{r}=""),"",D{r}-E{r})'.format(r=row))
        balance.number_format = MONEY
        for col in (1, 2, 3, 4, 5, 7, 8, 9):
            cell = ws.cell(row=row, column=col)
            cell.fill = INPUT_FILL
            cell.protection = UNLOCKED
        for col in range(1, len(headers) + 1):
            ws.cell(row=row, column=col).border = BOX
        ws.row_dimensions[row].height = 18

    waiver_dv = DataValidation(type="list", formula1="=WaiverStatus",
                               allow_blank=True, showDropDown=False)
    ws.add_data_validation(waiver_dv)
    waiver_dv.add("H%d:H%d" % (first, last))

    total_row = last + 1
    ws.cell(row=total_row, column=1, value="TOTALS").font = BOLD
    for col, letter in ((4, "D"), (5, "E"), (6, "F")):
        cell = ws.cell(row=total_row, column=col)
        cell.value = "=SUM({L}{f}:{L}{l})".format(L=letter, f=first, l=last)
        cell.number_format = MONEY
        cell.font = BOLD
    for col in range(1, len(headers) + 1):
        ws.cell(row=total_row, column=col).fill = TOTAL_FILL
        ws.cell(row=total_row, column=col).border = Border(
            left=THIN, right=THIN, top=MED, bottom=MED)

    checks = [
        "All work completed per contract",
        "Punch list items completed",
        "Final inspection passed",
        "Unconditional final lien waiver received",
    ]
    row = total_row + 2
    ws.cell(row=row, column=1,
            value="Work completion status before final payment:").font = BOLD
    row += 1
    for item in checks:
        ws.cell(row=row, column=1, value="☐  " + item).font = BODY
        row += 1

    footer_note(ws, row + 1, len(headers))
    set_widths(ws, [12, 28, 16, 14, 14, 14, 16, 20, 34])
    ws.print_title_rows = "%d:%d" % (header_row, header_row)

    return finish(wb, "7.4-payment-tracking.xlsx")


# ---------------------------------------------------------------------------
# 8.3  Material calculators
# ---------------------------------------------------------------------------


class CalcSheet:
    """Small builder for a single calculator sheet."""

    def __init__(self, wb, name, title, formula_line, first=False):
        self.ws = wb.active if first else wb.create_sheet(name)
        if first:
            self.ws.title = name
        self.ws.sheet_view.showGridLines = False
        title_block(self.ws, title, formula_line, span=4)
        set_widths(self.ws, [38, 16, 26, 30])
        self.row = 4

    def section(self, label):
        self.row += 1
        cell = self.ws.cell(row=self.row, column=1, value=label)
        cell.font = SECTION_FONT
        for col in range(1, 5):
            self.ws.cell(row=self.row, column=col).fill = HEADER_FILL
            self.ws.cell(row=self.row, column=col).border = HEADER_BORDER
        self.ws.row_dimensions[self.row].height = 20
        self.row += 1
        return self.row

    def input_cell(self, label, value, units, number_format=NUMBER):
        r = self.row
        self.ws.cell(row=r, column=1, value=label).font = BODY
        cell = self.ws.cell(row=r, column=2, value=value)
        cell.font = BODY
        cell.fill = INPUT_FILL
        cell.protection = UNLOCKED
        cell.number_format = number_format
        cell.alignment = Alignment(horizontal="right")
        cell.border = BOX
        self.ws.cell(row=r, column=3, value=units).font = NOTE_FONT
        self.row += 1
        return "B%d" % r

    def output_cell(self, label, formula, units, number_format=NUMBER,
                    bold=False):
        r = self.row
        cell = self.ws.cell(row=r, column=1, value=label)
        cell.font = BOLD if bold else BODY
        out = self.ws.cell(row=r, column=2, value=formula)
        out.font = BOLD if bold else BODY
        out.number_format = number_format
        out.alignment = Alignment(horizontal="right")
        out.border = BOX
        out.fill = SUBTOTAL_FILL if not bold else TOTAL_FILL
        self.ws.cell(row=r, column=3, value=units).font = NOTE_FONT
        self.row += 1
        return "B%d" % r

    def note(self, text):
        self.ws.cell(row=self.row, column=1, value=text).font = NOTE_FONT
        self.ws.merge_cells(start_row=self.row, start_column=1,
                            end_row=self.row, end_column=4)
        self.row += 1

    def blank(self, n=1):
        self.row += n

    def table(self, headers, rows, number_formats=None):
        start = self.row
        for col, text in enumerate(headers, start=1):
            self.ws.cell(row=start, column=col, value=text)
        style_header_row(self.ws, start, 1, len(headers), wrap=False)
        self.row += 1
        for values in rows:
            for col, value in enumerate(values, start=1):
                cell = self.ws.cell(row=self.row, column=col, value=value)
                cell.font = BODY
                cell.border = BOX
                if number_formats and col <= len(number_formats) and \
                        number_formats[col - 1]:
                    cell.number_format = number_formats[col - 1]
                    cell.alignment = Alignment(horizontal="right")
            self.row += 1
        end = self.row - 1
        self.row += 1
        return start, end

    def close(self):
        footer_note(self.ws, self.row + 1, 4)


def build_material_calculators():
    wb = Workbook()

    # -- Concrete ----------------------------------------------------------
    c = CalcSheet(wb, "Concrete", "CONCRETE CALCULATOR",
                  "Cubic yards = (Length ft × Width ft × Thickness in) ÷ 324",
                  first=True)
    c.section("INPUTS — type over the yellow cells")
    length = c.input_cell("Slab length", 20, "feet")
    width = c.input_cell("Slab width", 30, "feet")
    thickness = c.input_cell("Slab thickness", 4, "inches")
    waste = c.input_cell("Waste allowance", 0.05, "percent of volume", PCT)
    c.blank()
    c.section("RESULTS")
    area = c.output_cell("Slab area", "={L}*{W}".format(L=length, W=width),
                         "square feet", INT)
    cy = c.output_cell("Concrete required",
                       "=({L}*{W}*{T})/324".format(L=length, W=width,
                                                   T=thickness),
                       "cubic yards")
    with_waste = c.output_cell("With waste allowance",
                               "={C}*(1+{Wst})".format(C=cy, Wst=waste),
                               "cubic yards")
    c.output_cell("ORDER THIS MUCH", "=ROUNDUP({W},0)".format(W=with_waste),
                  "cubic yards (rounded up)", INT, bold=True)
    c.blank()
    c.section("REFERENCE — coverage by slab thickness")
    c.table(["Slab Thickness", "Square Feet per Cubic Yard"],
            [["4 inches", 81], ["5 inches", 65], ["6 inches", 54],
             ["8 inches", 40]],
            [None, INT])
    c.note("Ready-mix is sold in whole cubic yards. Running short mid-pour is "
           "far more expensive than ordering a spare yard.")
    c.close()

    # -- Drywall -----------------------------------------------------------
    d = CalcSheet(wb, "Drywall", "DRYWALL CALCULATOR",
                  "Sheets = Total area ÷ 32 sq ft (a 4' × 8' sheet)")
    d.section("INPUTS — type over the yellow cells")
    total_area = d.input_cell("Total wall and ceiling area", 800,
                              "square feet", INT)
    sheet_size = d.input_cell("Coverage per sheet", 32,
                              "square feet (4×8 = 32, 4×12 = 48)", INT)
    d_waste = d.input_cell("Waste allowance", 0.10, "percent", PCT)
    seams = d.input_cell("Linear feet of seams", 400, "linear feet", INT)
    d.blank()
    d.section("RESULTS")
    sheets = d.output_cell("Sheets needed",
                           "={A}/{S}".format(A=total_area, S=sheet_size),
                           "sheets")
    sheets_waste = d.output_cell("With waste allowance",
                                 "={S}*(1+{W})".format(S=sheets, W=d_waste),
                                 "sheets")
    order = d.output_cell("ORDER THIS MANY",
                          "=ROUNDUP({S},0)".format(S=sheets_waste),
                          "sheets (rounded up)", INT, bold=True)
    d.output_cell("Joint compound",
                  "=ROUNDUP({A}/100,0)".format(A=total_area),
                  "gallons (1 gal per 100 sq ft)", INT)
    d.output_cell("Joint tape", "=ROUNDUP({S}/100,0)".format(S=seams),
                  "rolls (1 roll per 100 linear ft)", INT)
    d.output_cell("Screws", "=ROUNDUP({O}/8,0)".format(O=order),
                  "pounds (1 lb per 8 sheets)", INT)
    d.blank()
    d.section("REFERENCE — sheet coverage")
    d.table(["Sheet Size", "Coverage (sq ft)"],
            [["4' × 8'", 32], ["4' × 10'", 40], ["4' × 12'", 48]],
            [None, INT])
    d.note("Use 4' × 8' sheets on walls and 4' × 12' on ceilings where you can "
           "handle them — fewer butt joints to finish.")
    d.close()

    # -- Paint -------------------------------------------------------------
    p = CalcSheet(wb, "Paint", "PAINT CALCULATOR",
                  "Gallons = (Total area ÷ coverage per gallon) × number of coats")
    p.section("INPUTS — type over the yellow cells")
    p_area = p.input_cell("Total area to paint", 1200, "square feet", INT)
    coverage = p.input_cell("Coverage per gallon", 350,
                            "square feet (see table below)", INT)
    coats = p.input_cell("Number of coats", 2, "coats", INT)
    p.blank()
    p.section("RESULTS")
    per_coat = p.output_cell("Gallons per coat",
                             "={A}/{C}".format(A=p_area, C=coverage), "gallons")
    total_gal = p.output_cell("Total gallons",
                              "={P}*{C}".format(P=per_coat, C=coats), "gallons")
    p.output_cell("BUY THIS MANY", "=ROUNDUP({T},0)".format(T=total_gal),
                  "gallons (rounded up)", INT, bold=True)
    p.blank()
    p.section("REFERENCE — coverage by surface")
    p.table(["Surface Type", "Coverage per Gallon (sq ft)"],
            [["Smooth drywall / plaster", 400],
             ["Textured drywall", 350],
             ["Rough wood / concrete", 300],
             ["Primer on new drywall", 300]],
            [None, INT])
    p.note("Buy all of one colour in a single trip so the batch numbers match.")
    p.close()

    # -- Flooring ----------------------------------------------------------
    f = CalcSheet(wb, "Flooring", "FLOORING CALCULATOR",
                  "Square feet to order = (Length × Width) × waste factor")
    f.section("INPUTS — type over the yellow cells")
    f_len = f.input_cell("Room length", 12, "feet")
    f_wid = f.input_cell("Room width", 15, "feet")
    f_factor = f.input_cell("Waste factor", 1.10,
                            "pick from the dropdown (1.10 = 10% waste)", '0.00')
    f.blank()
    f.section("RESULTS")
    f_area = f.output_cell("Floor area",
                           "={L}*{W}".format(L=f_len, W=f_wid),
                           "square feet", INT)
    f_total = f.output_cell("With waste factor",
                            "={A}*{F}".format(A=f_area, F=f_factor),
                            "square feet")
    f.output_cell("ORDER THIS MUCH", "=ROUNDUP({T},0)".format(T=f_total),
                  "square feet (round up to full boxes)", INT, bold=True)
    f.blank()
    f.section("REFERENCE — recommended waste factors")
    start, end = f.table(
        ["Flooring Type", "Waste Factor", "Waste %"],
        [["Carpet, sheet vinyl", 1.05, "5%"],
         ["Hardwood, straight pattern", 1.10, "10%"],
         ["Laminate, complex room", 1.15, "15%"],
         ["Tile, diagonal pattern", 1.20, "20%"]],
        [None, '0.00', None])
    factor_dv = DataValidation(type="list",
                               formula1="=$B$%d:$B$%d" % (start + 1, end),
                               allow_blank=False, showDropDown=False)
    f.ws.add_data_validation(factor_dv)
    factor_dv.add(f_factor)
    f.note("Buy every box in one order — dye lots vary between production runs.")
    f.close()

    # -- Roofing -----------------------------------------------------------
    r = CalcSheet(wb, "Roofing", "ROOFING CALCULATOR",
                  "Squares = (Footprint area × pitch multiplier) ÷ 100")
    r.section("INPUTS — type over the yellow cells")
    r_len = r.input_cell("Building footprint length", 40, "feet")
    r_wid = r.input_cell("Building footprint width", 30, "feet")
    pitch_cell_row = r.row
    r.ws.cell(row=pitch_cell_row, column=1, value="Roof pitch").font = BODY
    pitch = r.ws.cell(row=pitch_cell_row, column=2, value="6:12")
    pitch.font = BODY
    pitch.fill = INPUT_FILL
    pitch.protection = UNLOCKED
    pitch.alignment = Alignment(horizontal="right")
    pitch.border = BOX
    r.ws.cell(row=pitch_cell_row, column=3,
              value="rise : run — pick from the dropdown").font = NOTE_FONT
    r.row += 1
    pitch_ref = "B%d" % pitch_cell_row
    r_waste = r.input_cell("Waste allowance", 0.10, "percent", PCT)
    ridge = r.input_cell("Ridge and hip length", 60, "linear feet", INT)
    eave = r.input_cell("Eave length", 80, "linear feet", INT)
    r.blank()

    # The lookup table is laid out before the results block so the VLOOKUP and
    # the pitch dropdown can both point at a known range.
    r.section("PITCH MULTIPLIER LOOKUP")
    lookup_start, lookup_end = r.table(
        ["Roof Pitch", "Multiplier", "Example: 1,000 sq ft footprint"],
        [["Flat to 3:12", 1.03, "1,030 sq ft of roof"],
         ["4:12", 1.05, "1,050 sq ft of roof"],
         ["5:12", 1.08, "1,080 sq ft of roof"],
         ["6:12", 1.12, "1,120 sq ft of roof"],
         ["8:12", 1.20, "1,200 sq ft of roof"],
         ["10:12", 1.30, "1,300 sq ft of roof"],
         ["12:12", 1.41, "1,410 sq ft of roof"]],
        [None, '0.00', None])
    lookup_range = "$A$%d:$B$%d" % (lookup_start + 1, lookup_end)

    pitch_dv = DataValidation(
        type="list",
        formula1="=$A$%d:$A$%d" % (lookup_start + 1, lookup_end),
        allow_blank=False, showDropDown=False)
    pitch_dv.error = "Pick a pitch from the lookup table on this sheet."
    pitch_dv.errorTitle = "Unknown pitch"
    r.ws.add_data_validation(pitch_dv)
    pitch_dv.add(pitch_ref)

    r.section("RESULTS")
    footprint = r.output_cell("Footprint area",
                              "={L}*{W}".format(L=r_len, W=r_wid),
                              "square feet", INT)
    multiplier = r.output_cell(
        "Pitch multiplier",
        "=VLOOKUP({P},{R},2,FALSE)".format(P=pitch_ref, R=lookup_range),
        "from the lookup table above", '0.00')
    roof_area = r.output_cell("Roof surface area",
                              "={F}*{M}".format(F=footprint, M=multiplier),
                              "square feet", INT)
    squares = r.output_cell("Squares", "={A}/100".format(A=roof_area),
                            "squares (1 square = 100 sq ft)")
    sq_waste = r.output_cell("With waste allowance",
                             "={S}*(1+{W})".format(S=squares, W=r_waste),
                             "squares")
    order_sq = r.output_cell("ORDER THIS MANY",
                             "=ROUNDUP({S},0)".format(S=sq_waste),
                             "squares (rounded up)", INT, bold=True)
    r.output_cell("Ridge cap", "=ROUNDUP({R}/30,0)".format(R=ridge),
                  "bundles (1 per 30 linear ft)", INT)
    r.output_cell("Starter strip", "=ROUNDUP({E}/100,0)".format(E=eave),
                  "bundles (1 per 100 linear ft)", INT)
    r.output_cell("Underlayment", "=ROUNDUP({S}/4,0)".format(S=order_sq),
                  "rolls (4 squares per roll)", INT)
    r.blank()
    r.note("Footprint is the area the roof covers when viewed from above, not "
           "the sloped surface. The multiplier converts one to the other.")
    r.note("These are estimates. Verify measurements on site before ordering.")
    r.close()

    return finish(wb, "8.3-material-calculators.xlsx")


# ---------------------------------------------------------------------------
# Verification
# ---------------------------------------------------------------------------

EXPECTED = {
    "1.2-budget-tracking-spreadsheet.xlsx": {
        "sheets": ["Budget Tracker"],
        "needs": ["=SUM(", "-B", "=IF(AND(B"],
        "freeze": {"Budget Tracker": True},
    },
    "7.1-expense-tracking.xlsx": {
        "sheets": ["Expense Log", "Lists", "Category Summary"],
        "needs": ["=SUM(", "=SUMIF(", "=IF(F"],
        "freeze": {"Expense Log": True, "Category Summary": True},
    },
    "7.4-payment-tracking.xlsx": {
        "sheets": ["Payment Tracking", "Lists"],
        "needs": ["=SUM(", "=IF(AND(D"],
        "freeze": {"Payment Tracking": True},
    },
    "8.3-material-calculators.xlsx": {
        "sheets": ["Concrete", "Drywall", "Paint", "Flooring", "Roofing"],
        "needs": ["=VLOOKUP(", "=ROUNDUP(", "/324"],
        "freeze": {},
    },
}


def verify_xlsx(path):
    name = os.path.basename(path)
    spec = EXPECTED[name]
    problems = []
    wb = load_workbook(path)

    for sheet in spec["sheets"]:
        if sheet not in wb.sheetnames:
            problems.append("%s: missing sheet %r" % (name, sheet))

    formulas = []
    strings = []
    input_cells = 0
    for ws in wb.worksheets:
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str):
                    if cell.value.startswith("="):
                        formulas.append(cell.value)
                    else:
                        strings.append((ws.title, cell.coordinate, cell.value))
                if cell.fill is not None and cell.fill.fgColor is not None \
                        and cell.fill.fgColor.rgb in ("00FFF6CC", "FFFFF6CC"):
                    input_cells += 1

    joined = "\n".join(formulas)
    for needle in spec["needs"]:
        if needle not in joined:
            problems.append("%s: no formula containing %r" % (name, needle))

    for sheet, expected in spec["freeze"].items():
        if sheet in wb.sheetnames:
            frozen = wb[sheet].freeze_panes
            if expected and not frozen:
                problems.append("%s: %s header row is not frozen" % (name, sheet))

    for sheet, coord, value in strings:
        if "© 2024" in value or "(c) 2024" in value:
            problems.append("%s: stale copyright at %s!%s" % (name, sheet, coord))
        if "/Users/" in value or "file:///" in value:
            problems.append("%s: filesystem path at %s!%s" % (name, sheet, coord))
        if "____" in value:
            problems.append("%s: underscore run at %s!%s" % (name, sheet, coord))

    if not any(COPYRIGHT in v for _, _, v in strings):
        problems.append("%s: 2026 footer line missing" % name)

    validations = sum(len(ws.data_validations.dataValidation)
                      for ws in wb.worksheets)

    stats = {
        "sheets": len(wb.sheetnames),
        "formulas": len(formulas),
        "validations": validations,
        "input_cells": input_cells,
    }
    return problems, stats


def spot_check(path, checks):
    """Assert specific cells hold specific formulas."""
    problems = []
    wb = load_workbook(path)
    for sheet, coord, fragment in checks:
        value = wb[sheet][coord].value
        if not isinstance(value, str) or fragment not in value:
            problems.append("%s!%s expected %r, found %r"
                            % (sheet, coord, fragment, value))
    return problems


BUILDERS = [
    build_budget_tracker,
    build_expense_tracking,
    build_payment_tracking,
    build_material_calculators,
]


def main():
    os.makedirs(OUT_DIR, exist_ok=True)
    paths = [build() for build in BUILDERS]

    print("EXCEL WORKBOOKS")
    print("=" * 78)
    all_problems = []
    for path in paths:
        problems, stats = verify_xlsx(path)
        all_problems += problems
        status = "PASS" if not problems else "FAIL"
        print("  [%s] %-46s %7s bytes" % (status, os.path.basename(path),
                                          "{:,}".format(os.path.getsize(path))))
        print("         sheets=%(sheets)d  formulas=%(formulas)d  "
              "dropdowns=%(validations)d  input cells=%(input_cells)d" % stats)

    by_name = {os.path.basename(p): p for p in paths}
    spot = []
    spot += spot_check(by_name["1.2-budget-tracking-spreadsheet.xlsx"], [
        ("Budget Tracker", "D10", "C10-B10"),   # variance = Actual - Estimated
        ("Budget Tracker", "B14", "=SUM(B10:B13)"),  # first category subtotal
        ("Budget Tracker", "D14", "=C14-B14"),  # subtotal variance
    ])
    spot += spot_check(by_name["7.1-expense-tracking.xlsx"], [
        ("Expense Log", "G7", "SUM($F$7:F7)"),  # running total
        ("Expense Log", "F67", "=SUM(F7:F66)"),
        ("Category Summary", "C5", "SUMIF("),
    ])
    spot += spot_check(by_name["7.4-payment-tracking.xlsx"], [
        ("Payment Tracking", "F7", "D7-E7"),    # balance = due - paid
        ("Payment Tracking", "D47", "=SUM(D7:D46)"),
    ])
    spot += spot_check(by_name["8.3-material-calculators.xlsx"], [
        ("Concrete", "B14", "/324"),            # cubic yards
        ("Concrete", "B16", "=ROUNDUP("),       # order quantity
        ("Drywall", "B13", "/"),                # sheets = area / coverage
        ("Paint", "B12", "/"),                  # gallons per coat
        ("Flooring", "B13", "*"),               # area x waste factor
        ("Roofing", "B27", "VLOOKUP"),          # pitch multiplier
        ("Roofing", "B29", "/100"),             # squares
    ])
    all_problems += spot

    if all_problems:
        print("\n  PROBLEMS")
        for problem in all_problems:
            print("   - " + problem)
    else:
        print("\n  All checks passed: required sheets present, header rows")
        print("  frozen, =SUM / =SUMIF / variance / =VLOOKUP formulas verified")
        print("  by spot-check, no '© 2024', no underscore runs, no leaked paths.")
    print("  Output: %s" % OUT_DIR)
    return 1 if all_problems else 0


if __name__ == "__main__":
    raise SystemExit(main())
